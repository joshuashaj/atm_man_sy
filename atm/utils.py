import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

def save_to_excel(name, account_number, email, password):
    """Saves user data to an Excel file."""
    try:
        workbook = openpyxl.load_workbook('input_data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'User Input'
        headers = ['Name', 'Account Number', 'Email', 'Password']
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    except InvalidFileException:
        print("The file 'input_data.xlsx' is corrupted or invalid.")
        return

    sheet.append([name, account_number, email, password])
    workbook.save('input_data.xlsx')
    print("Data saved to input_data.xlsx")


def save_transaction_to_excel(user_name, account_number, transaction_type, amount, timestamp=None):
    """Saves transaction data to an Excel file."""
    try:
        workbook = openpyxl.load_workbook('transactions_data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Transactions'
        headers = ['User Name', 'Account Number', 'Transaction Type', 'Amount', 'Timestamp']
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    except InvalidFileException:
        print("The file 'transactions_data.xlsx' is corrupted or invalid.")
        return

    if not timestamp:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet.append([user_name, account_number, transaction_type, amount, timestamp])

    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in sheet[column]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save('transactions_data.xlsx')
    print("Transaction data saved to transactions_data.xlsx")

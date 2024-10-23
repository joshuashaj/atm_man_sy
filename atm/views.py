from django.shortcuts import render, redirect, get_object_or_404
from atm.models import BankAccountUser
from django.contrib import messages
import random, openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from django.contrib.auth.hashers import check_password, make_password


# Create your views here.
def openaccount(request):
    # This view handles the account opening process
    if request.method == 'POST':
        # Collecting form data
        name = request.POST.get('fullName')
        email = request.POST.get('email')
        phone_number = request.POST.get('phoneNumber')
        address = request.POST.get('address')
        account_type = request.POST.get('accountType')
        deposit = request.POST.get('deposit')
        terms = request.POST.get('terms') == 'on'  # Check if the terms are agreed upon
        password = f"{random.randint(0, 9999):04d}"  # Generate a random 4-digit password

        # Validate form data
        if all([terms, name, email, phone_number, address, account_type, deposit]):
            try:
                deposit = float(deposit)  # Ensure deposit is a valid number

                # Check if an account with the same email or phone number already exists
                if BankAccountUser.objects.filter(email=email).exists():
                    messages.error(request, 'An account with this email already exists.')
                    return redirect('openaccount')
                if BankAccountUser.objects.filter(phone_number=phone_number).exists():
                    messages.error(request, 'An account with this phone number already exists.')
                    return redirect('openaccount')

                # Hash the generated password for security
                password_hashed = make_password(password)

                # Create and save a new BankAccountUser instance
                new_account = BankAccountUser(
                    full_name=name,
                    email=email,
                    phone_number=phone_number,
                    address=address,
                    account_type=account_type,
                    deposit=deposit,
                    terms_agreed=terms,
                    password=password_hashed
                )
                new_account.save()

                # Save user details to an Excel file
                save_to_excel(name, new_account.account_number, email, password)

                # Display success message with account number and plain text password
                messages.success(request, f'Account successfully created!\n'
                                          f'Account Number: {new_account.account_number}\n'
                                          f'Password: {password}\n'
                                          f'NB: Save your password for further transactions')
                return redirect('success')  # Redirect to a success page
            except ValueError:
                # Handle invalid deposit value
                messages.error(request, 'Invalid deposit amount.')
        else:
            messages.error(request, 'Please fill in all required fields and agree to the terms.')

    return render(request, 'openaccount.html')


def save_to_excel(name, account_number, email, password):
    """Saves user data to an Excel file."""
    try:
        # Try to load the existing workbook
        workbook = openpyxl.load_workbook('input_data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook and sheet if it doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'User Input'

        # Add headers to the new sheet
        headers = ['Name', 'Account Number', 'Email', 'Password']
        sheet.append(headers)

        # Format headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)  # Make header bold
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background

    except InvalidFileException:
        print("The file 'input_data.xlsx' is corrupted or invalid.")
        return

    # Append the new user data to the next row
    sheet.append([name, account_number, email, password])

    # Save the workbook
    workbook.save('input_data.xlsx')
    print("Data saved to input_data.xlsx")


def success(request):
    return render(request, "success.html")


def login(request):
    # This view handles user login
    if request.method == 'POST':
        accountnumber = request.POST.get("accountnumber")
        password = request.POST.get("password")

        # Check if the user with the provided account number exists
        user = BankAccountUser.objects.filter(account_number=accountnumber).first()

        # Validate user credentials
        if user and check_password(password, user.password):
            request.session['accountnumber'] = accountnumber
            return redirect('home')  # Redirect to home page upon successful login
        elif user and not check_password(password, user.password):
            return render(request, "login.html", {'error': 'Incorrect password'})
        else:
            return render(request, "login.html", {'error': 'Account not found'})

    return render(request, "login.html")


def withdraw(request):
    # Placeholder for withdraw functionality
    return render(request, 'withdraw.html')


def home(request):
    # This view renders the home page for logged-in users
    if 'accountnumber' in request.session:
        accountnumber = request.session['accountnumber']
        user = BankAccountUser.objects.get(account_number=accountnumber)
        return render(request, 'Home.html', {'user': user})
    return redirect('login')


def logout(request):
    # This view handles user logout
    if 'accountnumber' in request.session:
        del request.session['accountnumber']
        request.session.flush()
    return redirect('login')


def changepassword(request, id):
    # Check if the user is logged in
    if 'accountnumber' in request.session:
        accountnumber = request.session['accountnumber']
        user = get_object_or_404(BankAccountUser, id=id)

        if request.method == 'POST':
            password = request.POST.get("password")  # Old password
            new_password = request.POST.get("new_password")
            re_password = request.POST.get("re_password")

            # Check if the old password is correct
            if check_password(password, user.password):
                # Ensure the new password and re-entered password match
                if new_password == re_password:
                    # Hash the new password before saving
                    user.password = make_password(new_password)
                    user.save()

                    # Update the Excel file with the new password
                    save_to_excel(user.full_name, user.account_number, user.email, new_password)

                    # Optionally, update session to reflect password change
                    request.session['accountnumber'] = accountnumber

                    # Redirect to home after successful password change
                    messages.success(request, "Password successfully changed.")
                    return redirect('../home')
                else:
                    messages.error(request, "New passwords do not match.")
            else:
                # Old password is incorrect
                messages.error(request, "Incorrect old password.")

        # If the request is not POST, render the change password form
        return render(request, 'change_password.html', {'user': user})

    # If the user is not logged in, redirect to login
    return redirect('login')

def forget_password(request):
    all_user = BankAccountUser.objects.all()
    if request.method == 'POST':
        account_number = request.POST.get("account_number")
        password = request.POST.get("generated_password")

        # Check if the account number exists
        if BankAccountUser.objects.filter(account_number=account_number).exists():
            user = BankAccountUser.objects.get(account_number=account_number)

            # Hash the new password before saving
            user.password = make_password(password)
            user.save()

            # Update the Excel file with the new password
            save_to_excel(user.full_name, user.account_number, user.email, password)

            # Optionally, display a success message
            messages.success(request,"Your password has been reset successfully. Please log in with your new password.")
            return redirect('login')  # Redirect to login page
        else:
            messages.error(request, "Account number does not exist. Please check and try again.")

    return render(request, "forgot_password.html",{"users":all_user})
